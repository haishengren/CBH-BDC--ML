import os
import openpyxl
import pandas as pd
from pycbh import smi2mh, smifromfn
from openpyxl.utils import get_column_letter


SMILES_file = r"...\SMILES_data.xlsx"  # Prepare a xlsx file with your SMILES data
txt_file = r'...\database\File.txt'  # Prepare an empty file in txt format, as the pycbh package accepts SMILES data in text format

df = pd.read_excel(SMILES_file)
df['formula smile'] = df[df.columns[:2]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
df[['formula smile']].to_csv(txt_file, index=False)

# Set up the CBH rung
rungs = [0]


def main():
    """
    USAGE: Generate CBH vectors and molecular fragment labels
    """
    smiles_arr = smifromfn(txt_file)
    print_messages = []

    for x in rungs:
        print(f'Calculating CBH-{x}')
        cbh_vector, good_x, labels = smi2mh(smiles_arr[1:], x)
        sparse_matrix_file = r'E:\监督学习\ymx_get_fragments_vectors\database\CBH-{} fragments matrix.txt'.format(x)

        with open(sparse_matrix_file, "w") as Y:
            Y.write(" ".join(map(str, labels)) + "\n")
            for idx, vector in enumerate(cbh_vector):
                Y.write(" ".join(map(str, vector)) + "\n")

        print_messages.append(
            f'\nThe CBH-{x} fragments matrix has been generated\n'
            f'Number of species: {len(good_x)}\n'
            f'The length of the input vector: {len(labels)}'
        )
    print('\n'.join(print_messages))


if __name__ == '__main__':
    main()


# Standardize Ensemble learning input file formats
def split_data_to_columns(file_path, start_column):
    with open(file_path, 'r') as file:
        data_lines = file.readlines()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Get the column index of the starting column
    start_column_index = ord(start_column) - ord('A') + 1

    # Loop processing of each row of data
    for row_index, line in enumerate(data_lines, start=1):
        data = line.strip().split()  # Split data by space
        # Place the segmented data one by one in the corresponding columns
        for i, value in enumerate(data):
            column_index = start_column_index + i
            column_letter = get_column_letter(column_index)
            sheet[f"{column_letter}{row_index}"] = value

    # Save Workbook to File
    excel_file_path = file_path.replace('.txt', '.xlsx')
    workbook.save(excel_file_path)
    os.remove(file_path)


for rung in rungs:
    input_file = r'E:\监督学习\ymx_get_fragments_vectors\database\CBH-{} fragments matrix.txt'.format(rung)
    split_data_to_columns(input_file, 'A')
